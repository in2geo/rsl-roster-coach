#!/usr/bin/env python3
"""Build the tag-review WORKSHEET from the JSON bundle.

WHY A SHEET: this project's tag pipeline has always been worksheet-based
(DB_Champion_Tags -> seeds -> live). A sheet gives the reviewer somewhere to WRITE
VERDICTS BACK, which JSON does not, and it closes policy #18 ("any tag ruling MUST be
written back to the worksheet in the SAME session or it is ORPHANED") — the exact
failure that let the hand-analysis in seeds 42/43 be silently overwritten.

Run tools/export-tag-review.mjs first (it produces the JSON this reads).

    node --env-file=.env.local tools/export-tag-review.mjs
    python tools/export-tag-review-xlsx.py

Emits output/tag-review/TAG_REVIEW.xlsx with:
    Tag_Review   one row per champion_tag, WITH the verbatim skill text alongside,
                 plus empty VERDICT / EVIDENCE / POLICY / NOTES columns to fill in
    Skills       one row per skill — verbatim Plarium text (the ground truth)
    No_Evidence  champions with NO skill text: cannot be reviewed, do not infer
    Vocabulary   the controlled tag vocabulary
    Policies     the Tag Review Policies, verbatim from CLAUDE.md
    README       the brief

Contains no credentials.
"""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
SRC = os.path.join(REPO, 'output', 'tag-review')
OUT = os.path.join(SRC, 'TAG_REVIEW.xlsx')

if not os.path.exists(os.path.join(SRC, 'champions.json')):
    sys.exit(f"missing {SRC}/champions.json — run: node --env-file=.env.local tools/export-tag-review.mjs")

champs = json.load(open(os.path.join(SRC, 'champions.json'), encoding='utf-8'))
vocab = json.load(open(os.path.join(SRC, 'vocabulary.json'), encoding='utf-8'))
policies = open(os.path.join(SRC, 'POLICIES.md'), encoding='utf-8').read()

HDR = Font(bold=True, color='FFFFFF')
HDRFILL = PatternFill('solid', fgColor='2F5597')
FILLME = PatternFill('solid', fgColor='FFF2CC')   # columns the reviewer fills in
TOP = Alignment(vertical='top', wrap_text=True)

wb = Workbook()

def header(ws, cols):
    ws.append([c[0] for c in cols])
    for i, (name, width) in enumerate(cols, start=1):
        ws.cell(row=1, column=i).font = HDR
        ws.cell(row=1, column=i).fill = HDRFILL
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = 'A2'

# ---------------- Tag_Review: the working sheet ----------------
ws = wb.active
ws.title = 'Tag_Review'
cols = [('Champion',24), ('Rarity',11), ('Faction',18), ('Affinity',9),
        ('Tag',24), ('Status',10), ('Source_Type',18), ('Source_Note',34),
        ('Skill_Text (VERBATIM — the ground truth)',95),
        ('VERDICT',12), ('EVIDENCE (verbatim quote)',48), ('POLICY',10), ('NOTES',28)]
header(ws, cols)

rows = 0
for c in champs:
    if not c['current_tags']:
        continue
    blob = '\n\n'.join(f"[{s['slot']}] {s['name']}: {s['text']}"
                       for s in c['skills'] if s.get('text'))
    for t in sorted(c['current_tags'], key=lambda x: (x['status'], x['tag'])):
        ws.append([c['name'], c['rarity'], c['faction'], c['affinity'],
                   t['tag'], t['status'], t.get('source_type') or '', t.get('note') or '',
                   blob, '', '', '', ''])
        rows += 1

for r in range(2, rows + 2):
    for col in (9, 11):
        ws.cell(row=r, column=col).alignment = TOP
    for col in (10, 11, 12, 13):
        ws.cell(row=r, column=col).fill = FILLME

dv = DataValidation(type='list', formula1='"keep,reject,add,no_evidence,unsure"', allow_blank=True)
dv.error = 'Use: keep / reject / add / no_evidence / unsure'
ws.add_data_validation(dv)
dv.add(f'J2:J{rows+1}')
ws.auto_filter.ref = f'A1:M{rows+1}'

# ---------------- Skills ----------------
ws2 = wb.create_sheet('Skills')
header(ws2, [('Champion',24), ('Rarity',11), ('Slot',8), ('Skill_Name',28),
             ('Skill_Text (VERBATIM)',110), ('Cooldown',10), ('Asc_Req',8)])
n2 = 0
for c in champs:
    for s in c['skills']:
        ws2.append([c['name'], c['rarity'], s['slot'], s['name'], s.get('text') or '',
                    s.get('cooldown') or '', s.get('ascension_required')])
        n2 += 1
for r in range(2, n2 + 2):
    ws2.cell(row=r, column=5).alignment = TOP
ws2.auto_filter.ref = f'A1:G{n2+1}'

# ---------------- No_Evidence ----------------
ws3 = wb.create_sheet('No_Evidence')
header(ws3, [('Champion',26), ('Rarity',11), ('Faction',18), ('Current_Tags',60)])
noev = [c for c in champs if not any(s.get('text') for s in c['skills'])]
for c in noev:
    ws3.append([c['name'], c['rarity'], c['faction'],
                ', '.join(t['tag'] for t in c['current_tags'])])
ws3['F1'] = 'These champions have NO skill text. They CANNOT be reviewed — mark no_evidence, never infer a tag.'
ws3['F1'].font = Font(bold=True, color='C00000')

# ---------------- Vocabulary ----------------
ws4 = wb.create_sheet('Vocabulary')
header(ws4, [('Tag',30), ('Description',110)])
for v in sorted(vocab, key=lambda x: x['name']):
    ws4.append([v['name'], v.get('description') or ''])
for r in range(2, len(vocab) + 2):
    ws4.cell(row=r, column=2).alignment = TOP

# ---------------- Policies ----------------
ws5 = wb.create_sheet('Policies')
ws5.column_dimensions['A'].width = 150
ws5['A1'] = 'Tag Review Policies — verbatim from CLAUDE.md. THESE ARE THE REVIEW CRITERIA.'
ws5['A1'].font = Font(bold=True, size=12, color='C00000')
for i, line in enumerate(policies.split('\n'), start=3):
    ws5.cell(row=i, column=1, value=line).alignment = Alignment(wrap_text=True, vertical='top')

# ---------------- README ----------------
ws6 = wb.create_sheet('README', 0)
ws6.column_dimensions['A'].width = 118
readme = f"""TAG REVIEW WORKSHEET
Snapshot: {__import__('datetime').date.today().isoformat()}.  Read-only data. No credentials in this file.

WHAT IS HERE
  Tag_Review   {rows} rows — one per champion_tag, with that champion's VERBATIM skill text alongside.
               Fill in the four shaded columns: VERDICT / EVIDENCE / POLICY / NOTES.
  Skills       {n2} skills — verbatim Plarium text.
  No_Evidence  {len(noev)} champions with NO skill text. Cannot be reviewed. Mark no_evidence; never infer.
  Vocabulary   {len(vocab)} tags in the controlled vocabulary.
  Policies     The Tag Review Policies, verbatim from CLAUDE.md — the review criteria. READ FIRST.

THE JOB
  For each row, decide whether the Tag is what the Skill_Text actually supports, per Policies.
  Skill_Text is verbatim Plarium text and is the GROUND TRUTH. Judge against the text — never
  against a tier list, a guide, or prior belief.

  VERDICT   keep | reject | add | no_evidence | unsure   (dropdown)
  EVIDENCE  the VERBATIM clause from Skill_Text that decides it. REQUIRED.
            A verdict with no quote will not be actioned — that is the rule that keeps this
            reviewable, and it is how the 7 known false positives were caught.
  POLICY    the policy number you applied, e.g. #16, #19, #20
  NOTES     anything ambiguous. Prefer 'unsure' over a guess: being wrong costs more than
            being uncertain.

WHAT USUALLY GOES WRONG
  The recurring error is a bracket scraper reading a NON-PLACEMENT clause as a placement.
  A [Bracket] is NOT placed when it follows:
    ignore ............ "ignores [Shield] buffs"                                    (#16)
    remove/strip/steal  "removes all [Increase DEF]" -> earns Buff Strip, not Increase DEF (#19)
    instantly activates "instantly activates [Poison]" -> Debuff Activation          (#12)
    immune to (#10) · increases the duration of (#11) · transfers/redirects (#13) · except (#14)
    a self-condition .. "while this Champion is under a [Veil] buff" — a PREREQUISITE she must
                        receive from elsewhere, not a buff she places.  (proposed #20)

  LOAD-BEARING DISTINCTIONS
    [Veil] != [Perfect Veil] — different buffs, both in the vocabulary. Three champions place
      Perfect Veil and were wrongly tagged Veil off a condition clause.
    #17 vs #20 — "this debuff cannot be resisted if this Champion is under [Veil]": the DEBUFF
      is still placed, so tag it (#17 APPROVE); but [Veil] itself is not hers (#20 REJECT).
      Both halves of one sentence.
    Self-combo (#1 exception) — if the SAME champion places the prerequisite debuff herself,
      she delivers the chain unaided -> APPROVE (Frozen Banshee, Coldheart).

KNOWN-BAD ROWS ALREADY FOUND (2026-07-17) — expect to rediscover these; they validate the method:
    Rhaia     Veil, Perfect Veil  -- places no veil at all; passive REQUIRES her to be under one
    Umetogi   Veil                -- places [Perfect Veil] on A3, never [Veil]
    Yannica   Veil, Shield        -- places [Perfect Veil]; Shield is ignored (#16) + removed (#19)
    Yumeko    Veil                -- passive places [Perfect Veil], never [Veil]
    Elegaius  Heal Reduction      -- negated self-condition
  These are already written up in seeds/166_self_condition_tag_corrections_PROPOSED.sql (not applied).

WHAT HAPPENS NEXT
  Findings -> a committed seeds/*.sql -> human approval -> live. No auto-merge (CLAUDE.md).
  Policy #18: any tag ruling MUST be written back to the worksheet in the same session, or it is
  ORPHANED. This sheet IS that writeback. Return the filled-in file.
"""
for i, line in enumerate(readme.split('\n'), start=1):
    ws6.cell(row=i, column=1, value=line).alignment = Alignment(wrap_text=False, vertical='top')
ws6['A1'].font = Font(bold=True, size=14)

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  Tag_Review   {rows} tag rows to review")
print(f"  Skills       {n2} skills (verbatim)")
print(f"  No_Evidence  {len(noev)} champions with no skill text")
print(f"  Vocabulary   {len(vocab)} tags")
print(f"  Policies     verbatim from CLAUDE.md")
